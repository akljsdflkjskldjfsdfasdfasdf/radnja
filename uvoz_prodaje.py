# uvoz_prodaje.py — čitanje fajla sa prodajom (CSV ili Excel) i upis u bazu.
# Fajl mora imati zaglavlje sa kolonama: barkod i kolicina; datum može biti
# kolona u fajlu ILI se bira na ekranu (BizniSoft izveštaji često nemaju datum po redu).
# Gutamo razne varijante: ";" ili ",", datume "2026-07-18" i "18.07.2026",
# količine "3" i "2,5", stara Windows slova, barkod upisan kao broj u Excelu.

import csv
import io
from datetime import date, datetime

# Dozvoljena imena kolona u fajlu -> naše ime (mala slova).
# Uključena i verovatna imena iz BizniSoft izveštaja.
IMENA_KOLONA = {
    "datum": "datum", "dan": "datum",
    "barkod": "barkod", "ean": "barkod", "bar-kod": "barkod", "bar kod": "barkod",
    "naziv": "naziv", "artikal": "naziv", "proizvod": "naziv",
    "naziv artikla": "naziv", "naziv proizvoda": "naziv",
    "kolicina": "kolicina", "količina": "kolicina", "kom": "kolicina", "qty": "kolicina",
    "prodata kolicina": "kolicina", "prodata količina": "kolicina", "izlaz": "kolicina",
}

FORMATI_DATUMA = ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%Y.")


def _u_tekst(sadrzaj_bajtova):
    """Pretvori bajtove fajla u tekst (novi programi: UTF-8; stari: Windows-1250)."""
    try:
        return sadrzaj_bajtova.decode("utf-8-sig")
    except UnicodeDecodeError:
        return sadrzaj_bajtova.decode("cp1250")


def _u_datum(tekst):
    """Prihvati "2026-07-18" ili "18.07.2026" -> uvek vrati "2026-07-18"."""
    for oblik in FORMATI_DATUMA:
        try:
            return datetime.strptime((tekst or "").strip(), oblik).date().isoformat()
        except ValueError:
            pass
    return None


def _celija_u_tekst(vrednost):
    """Excel ćeliju pretvori u tekst: datume u GGGG-MM-DD, barkod-broj bez ".0"."""
    if vrednost is None:
        return ""
    if isinstance(vrednost, datetime):
        return vrednost.date().isoformat()
    if isinstance(vrednost, date):
        return vrednost.isoformat()
    if isinstance(vrednost, float) and vrednost == int(vrednost):
        return str(int(vrednost))      # Excel barkod često čuva kao broj (8.6e+12)
    return str(vrednost).strip()


def _lepa_kolicina(broj):
    """3.0 prikaži kao 3, a 2.5 ostavi kao 2.5."""
    return int(broj) if broj == int(broj) else broj


def _redovi_iz_csv(sadrzaj_bajtova):
    """CSV fajl -> (imena kolona, lista redova-rečnika)."""
    tekst = _u_tekst(sadrzaj_bajtova)
    linije = tekst.splitlines()
    if not linije:
        return [], []
    # Srpski Excel često koristi ";" umesto "," — pogodi po prvom redu.
    razdvajac = ";" if linije[0].count(";") >= linije[0].count(",") else ","
    citac = csv.DictReader(io.StringIO(tekst), delimiter=razdvajac)
    return list(citac.fieldnames or []), [dict(red) for red in citac]


def _redovi_iz_xlsx(sadrzaj_bajtova):
    """Excel (.xlsx) fajl -> (imena kolona, lista redova-rečnika). Čita prvi list."""
    from openpyxl import load_workbook
    knjiga = load_workbook(io.BytesIO(sadrzaj_bajtova), read_only=True, data_only=True)
    redovi = knjiga.active.iter_rows(values_only=True)
    zaglavlje = next(redovi, None)
    if not zaglavlje:
        return [], []
    imena = [str(c).strip() if c is not None else "" for c in zaglavlje]
    lista = [
        {ime: _celija_u_tekst(vrednost) for ime, vrednost in zip(imena, red)}
        for red in redovi
    ]
    knjiga.close()
    return imena, lista


def uvezi_fajl(sadrzaj_bajtova, ime_fajla, store_id, veza, podrazumevani_datum=None):
    """Pročitaj CSV ili Excel fajl i upiši prodaju. Vraća izveštaj šta se desilo.

    podrazumevani_datum se koristi za SVE redove ako fajl nema kolonu "datum"
    (tipično za BizniSoft izveštaj "promet po artiklima" za jedan dan).
    """
    je_excel = (ime_fajla or "").lower().endswith((".xlsx", ".xlsm")) \
        or sadrzaj_bajtova[:2] == b"PK"
    if je_excel:
        imena, redovi_fajla = _redovi_iz_xlsx(sadrzaj_bajtova)
    else:
        imena, redovi_fajla = _redovi_iz_csv(sadrzaj_bajtova)

    if not imena:
        return {"ok": False, "poruka_greske": "Fajl je prazan."}

    # Prevedi imena kolona iz fajla u naša ("Naziv artikla" -> "naziv").
    kolone = {}
    for ime in imena:
        nase = IMENA_KOLONA.get((ime or "").strip().lower())
        if nase and nase not in kolone:
            kolone[nase] = ime

    nedostaju = [k for k in ("barkod", "kolicina") if k not in kolone]
    if "datum" not in kolone and not podrazumevani_datum:
        nedostaju.append("datum (ili izaberi dan prodaje na ekranu)")
    if nedostaju:
        return {"ok": False, "poruka_greske": (
            f"U fajlu ne postoje kolone: {', '.join(nedostaju)}. "
            f"Nađeno u fajlu: {', '.join(n for n in imena if n)}."
        )}

    # Saberi po (barkod, dan) — ako fajl ima više redova za isti artikal, spoji ih.
    zbir = {}
    nazivi = {}
    greske = []
    for broj_reda, red in enumerate(redovi_fajla, start=2):   # red 1 je zaglavlje
        barkod = (red.get(kolone["barkod"]) or "").strip()
        if not barkod:
            continue                                          # prazan red preskačemo ćutke
        if "datum" in kolone:
            datum = _u_datum(red.get(kolone["datum"]))
            if datum is None:
                greske.append(f"red {broj_reda}: ne razumem datum '{red.get(kolone['datum'])}'")
                continue
        else:
            datum = podrazumevani_datum
        try:
            kolicina = float((red.get(kolone["kolicina"]) or "").strip().replace(",", "."))
        except ValueError:
            greske.append(f"red {broj_reda}: ne razumem količinu '{red.get(kolone['kolicina'])}'")
            continue
        zbir[(barkod, datum)] = zbir.get((barkod, datum), 0) + kolicina
        if "naziv" in kolone and red.get(kolone["naziv"]):
            nazivi[barkod] = red[kolone["naziv"]].strip()

    # Upis u bazu: poznate artikle upiši (isti dan = zameni), nepoznate prijavi.
    uvezeno = 0
    dani = set()
    nepoznati = {}
    for (barkod, datum), kolicina in zbir.items():
        proizvod = veza.execute(
            "SELECT id FROM proizvodi WHERE barkod = ?", (barkod,)
        ).fetchone()
        if proizvod is None:
            n = nepoznati.setdefault(barkod, {"naziv": nazivi.get(barkod, ""), "kolicina": 0})
            n["kolicina"] += kolicina
            continue
        veza.execute(
            """
            INSERT INTO prodaja (store_id, proizvod_id, datum, kolicina)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(store_id, proizvod_id, datum)
            DO UPDATE SET kolicina = excluded.kolicina
            """,
            (store_id, proizvod["id"], datum, kolicina),
        )
        uvezeno += 1
        dani.add(datum)
    veza.commit()

    return {
        "ok": True,
        "uvezeno": uvezeno,
        "dani": sorted(dani),
        "nepoznati": [
            {"barkod": b, "naziv": p["naziv"], "kolicina": _lepa_kolicina(p["kolicina"])}
            for b, p in nepoznati.items()
        ],
        "greske": greske,
    }
